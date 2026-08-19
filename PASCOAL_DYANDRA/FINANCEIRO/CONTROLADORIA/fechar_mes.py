"""
=============================================================================
  FECHAMENTO MENSAL A PARTIR DA PLANILHA RECEITA ESC 2026
  Pascoal & Dyandra Advocacia
=============================================================================

  Le a planilha do escritorio (export .xlsx do Google Sheets), apura o mes e
  gera um arquivo de fechamento com:

    RESUMO      - receita, despesa e resultado, com as lacunas explicitas
    RECEITA     - linha a linha, com as inconsistencias marcadas
    SAIDAS      - despesas do mes, separando pago / a pagar / sem valor
    PENDENCIAS  - o que precisa ser resolvido para o mes fechar

  A planilha de origem NAO e alterada. Ver docs/POP_FINANCEIRO.md.

  Uso:
    .venv\\Scripts\\python.exe FINANCEIRO/CONTROLADORIA/fechar_mes.py \\
        "RECEITA ESC 2026.xlsx" AGO26 --saidas "SAIDAS Ago26" --hoje 2026-08-19
=============================================================================
"""
import sys
import argparse
import unicodedata
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

# ---------------------------------------------------------------- estilos
AZUL = '1F4E79'
F_HDR = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
P_HDR = PatternFill('solid', start_color=AZUL, end_color=AZUL)
F_TIT = Font(name='Calibri', bold=True, size=14, color=AZUL)
F_BOLD = Font(name='Calibri', bold=True, size=11)
F_NORM = Font(name='Calibri', size=10)
F_ALERTA = Font(name='Calibri', size=10, bold=True, color='A2261C')
P_ALERTA = PatternFill('solid', start_color='FAE8E5', end_color='FAE8E5')
P_FALTA = PatternFill('solid', start_color='FFF3CD', end_color='FFF3CD')
P_OK = PatternFill('solid', start_color='E1EFE9', end_color='E1EFE9')
MOEDA = 'R$ #,##0.00'
BORDA = Border(*[Side('thin')] * 4)


def sem_acento(txt):
    txt = str(txt or '')
    return ''.join(c for c in unicodedata.normalize('NFD', txt)
                   if unicodedata.category(c) != 'Mn').upper().strip()


def num(v):
    return v if isinstance(v, (int, float)) else None


def achar_aba(wb, alvo):
    """Localiza a aba ignorando acento, caixa e espaco sobrando."""
    a = sem_acento(alvo).replace(' ', '')
    for nome in wb.sheetnames:
        if sem_acento(nome).replace(' ', '') == a:
            return wb[nome]
    raise SystemExit(f'Aba nao encontrada: {alvo}\nDisponiveis: {wb.sheetnames}')


# ============================================================
# LEITURA DA ABA DE RECEITA
# ============================================================

COLS = {'pasta': 1, 'cliente': 2, 'empresa': 3, 'bruto': 4, 'principal': 5,
        'sucumb': 6, 'hon_contrato': 7, 'hon_liq': 8, 'com_i': 9, 'com_j': 10,
        'repasse': 11, 'venc': 12, 'parc_ini': 13, 'parc_fim': 14,
        'parc_no': 15, 'dt_pagto': 16, 'dt_repasse': 17, 'obs': 18, 'obs2': 19}


def ler_receita(ws):
    """Le os blocos da aba (trabalhista no topo, INSS abaixo). Cada bloco tem
    cabecalho proprio comecando por 'Pasta' na coluna A."""
    cabecalhos = [r for r in range(1, ws.max_row + 1)
                  if sem_acento(ws.cell(r, 1).value) == 'PASTA']
    if not cabecalhos:
        raise SystemExit('Nenhum cabecalho "Pasta" encontrado na aba de receita.')

    blocos = []
    for i, h in enumerate(cabecalhos):
        limite = cabecalhos[i + 1] if i + 1 < len(cabecalhos) else ws.max_row + 1
        linhas, total_row = [], None
        for r in range(h + 1, limite):
            pasta = ws.cell(r, COLS['pasta']).value
            if pasta is None or str(pasta).strip() == '':
                # linha sem pasta mas com valor = linha de total do bloco
                if any(num(ws.cell(r, c).value) for c in (4, 8)):
                    total_row = r
                continue
            reg = {k: ws.cell(r, c).value for k, c in COLS.items()}
            reg['linha'] = r
            # o numero da pasta vem como float do Sheets (2098.0) - normalizar
            reg['pasta'] = (str(int(pasta)) if isinstance(pasta, float) and pasta.is_integer()
                            else str(pasta).strip())
            linhas.append(reg)
        blocos.append({'header': h, 'linhas': linhas, 'total_row': total_row,
                       'rotulo': 'TRABALHISTA' if i == 0 else 'PREVIDENCIARIO (INSS)'})
    return blocos


# Identidade contabil da planilha, verificada em 19/08/2026 sobre a aba AGO26:
# fecha ao centavo em 40 das 46 linhas com repasse, e as 6 que quebram sao
# justamente os lancamentos furados.
#
#     Valor Bruto = Repasse + Hon. Liquido + Comissao(J) - Comissao(I)
#
# Ou seja: a comissao da coluna J sai do bolso do escritorio, e a da coluna I
# entra. Confirmar com o escritorio o que a coluna I representa.
TOLERANCIA = 1.00     # centavos de arredondamento das parcelas


def conferir(blocos):
    """Aponta as inconsistencias linha a linha e as pastas duplicadas."""
    todas = [l for b in blocos for l in b['linhas']]

    vistos = {}
    for l in todas:
        vistos.setdefault(l['pasta'], []).append(l)

    for l in todas:
        graves, avisos = [], []
        bruto = num(l['bruto']) or 0
        hon = num(l['hon_liq'])
        rep = num(l['repasse'])
        ci = num(l['com_i']) or 0
        cj = num(l['com_j']) or 0

        # --- a conferencia principal: o rateio do bruto fecha?
        if bruto and rep is not None:
            dif = bruto - ((rep or 0) + (hon or 0) + cj - ci)
            l['diferenca'] = dif
            if abs(dif) > TOLERANCIA:
                if hon is None:
                    graves.append(f'honorario em branco - faltam R$ {dif:,.2f} do bruto')
                elif dif > 0:
                    graves.append(f'rateio nao fecha: R$ {dif:,.2f} do bruto sem destino')
                else:
                    graves.append(f'rateio nao fecha: R$ {-dif:,.2f} a mais que o bruto')

        if hon is not None and hon < 0:
            graves.append(f'honorario NEGATIVO (R$ {hon:,.2f}) - o caso deu prejuizo')
        if 'FULANO' in sem_acento(l['cliente']) or 'TESTE' in sem_acento(l['cliente']):
            graves.append('registro de teste somando no faturamento')
        if len(vistos[l['pasta']]) > 1:
            outros = [o['cliente'] for o in vistos[l['pasta']] if o is not l]
            graves.append(f'pasta duplicada (tambem em: {", ".join(str(o)[:20] for o in outros)})')

        if sem_acento(l['dt_repasse']) == 'X':
            avisos.append('repasse marcado "x" - nao realizado')
        if bruto and rep is None and 'HONORARIO' not in sem_acento(l['obs2']):
            avisos.append('sem repasse lancado')
        if not l['empresa']:
            avisos.append('sem parte contraria')
        # so vale como aviso quando NADA foi registrado: nem pagamento, nem repasse.
        # (a coluna OBS "ok" nao e preenchida em todas as linhas - nao serve de sinal)
        if l['dt_pagto'] is None and not str(l['dt_repasse'] or '').strip():
            avisos.append('sem data de pagamento e sem data de repasse')

        l['graves'], l['avisos'] = graves, avisos
        l['problemas'] = graves + avisos
        l['eh_teste'] = any('registro de teste' in p for p in graves)
    return todas


# ============================================================
# LEITURA DA ABA DE SAIDAS
# ============================================================

def ler_saidas(ws, ano, mes):
    """Classifica cada linha de despesa: paga no mes, a pagar, de outro mes
    (copiada) ou sem valor lancado."""
    itens, categoria = [], ''
    for r in range(2, ws.max_row + 1):
        desc = ws.cell(r, 2).value
        valor = num(ws.cell(r, 3).value)
        venc = ws.cell(r, 4).value
        pagto = ws.cell(r, 5).value

        # Categoria e sempre negrito e sem valor. Item nunca e as duas coisas -
        # ha item em negrito (com valor) e ha item sem valor (nao negrito).
        if desc and valor is None and ws.cell(r, 2).font.bold:
            categoria = str(desc).strip()
            continue
        if not desc:
            continue                                # vazia ou linha de total do bloco

        def dt(v):
            return v.date() if isinstance(v, datetime) else (v if isinstance(v, date) else None)

        dv, dp = dt(venc), dt(pagto)
        if valor is None:
            situacao = 'SEM VALOR LANCADO'
        elif dp and (dp.year, dp.month) == (ano, mes):
            situacao = 'PAGO NO MES'
        elif dp:
            situacao = f'PAGO EM {dp.strftime("%m/%Y")} (linha de outro mes)'
        elif dv:
            situacao = 'A PAGAR'
        else:
            situacao = 'SEM DATA'

        itens.append({'linha': r, 'categoria': categoria, 'desc': str(desc).strip(),
                      'valor': valor, 'venc': dv, 'pagto': dp, 'situacao': situacao})
    return itens


# ============================================================
# GERACAO DO FECHAMENTO
# ============================================================

def escrever(ws, linha, valores, negrito=False, moeda_cols=(), fill=None):
    for i, v in enumerate(valores, 1):
        c = ws.cell(linha, i, v)
        c.font = F_BOLD if negrito else F_NORM
        c.border = BORDA
        if i in moeda_cols:
            c.number_format = MOEDA
        if fill:
            c.fill = fill
    return linha + 1


def cabecalho(ws, linha, titulos, larguras=None):
    for i, t in enumerate(titulos, 1):
        c = ws.cell(linha, i, t)
        c.font, c.fill, c.border = F_HDR, P_HDR, BORDA
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if larguras:
        for i, w in enumerate(larguras, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 28
    return linha + 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('planilha')
    ap.add_argument('aba_receita')
    ap.add_argument('--saidas', required=True)
    ap.add_argument('--hoje', default=date.today().isoformat())
    ap.add_argument('--saida', default='FECHAMENTO.xlsx')
    args = ap.parse_args()

    hoje = date.fromisoformat(args.hoje)
    wb = openpyxl.load_workbook(args.planilha, data_only=True)

    ws_rec = achar_aba(wb, args.aba_receita)
    ws_sai = achar_aba(wb, args.saidas)

    blocos = ler_receita(ws_rec)
    todas = conferir(blocos)
    saidas = ler_saidas(ws_sai, hoje.year, hoje.month)

    # -------- apuracao da receita
    tot_bruto = sum(num(l['bruto']) or 0 for l in todas)
    tot_hon = sum(num(l['hon_liq']) or 0 for l in todas)
    tot_com = sum((num(l['com_i']) or 0) + (num(l['com_j']) or 0) for l in todas)
    tot_rep = sum(num(l['repasse']) or 0 for l in todas)
    hon_teste = sum(num(l['hon_liq']) or 0 for l in todas if l['eh_teste'])

    # -------- apuracao das saidas
    pago = sum(i['valor'] for i in saidas if i['situacao'] == 'PAGO NO MES')
    apagar = [i for i in saidas if i['situacao'] == 'A PAGAR']
    vencidas = [i for i in apagar if i['venc'] and i['venc'] < hoje]
    outro_mes = [i for i in saidas if i['situacao'].startswith('PAGO EM')]
    sem_valor = [i for i in saidas if i['situacao'] == 'SEM VALOR LANCADO']

    print('=' * 68)
    print(f'  FECHAMENTO {args.aba_receita} - posicao em {hoje.strftime("%d/%m/%Y")}')
    print('=' * 68)
    print(f'  Processos ................. {len(todas)}')
    print(f'  Valor bruto movimentado ... R$ {tot_bruto:>14,.2f}')
    print(f'  Honorario liquido ......... R$ {tot_hon:>14,.2f}')
    print(f'  (-) registro de teste ..... R$ {hon_teste:>14,.2f}')
    print(f'  = RECEITA REAL ............ R$ {tot_hon - hon_teste:>14,.2f}')
    print(f'  Comissoes ................. R$ {tot_com:>14,.2f}')
    print(f'  Repasses a cliente ........ R$ {tot_rep:>14,.2f}')
    print('-' * 68)
    print(f'  Despesa PAGA no mes ....... R$ {pago:>14,.2f}')
    print(f'  Despesa A PAGAR ........... R$ {sum(i["valor"] for i in apagar):>14,.2f}'
          f'   ({len(vencidas)} vencidas)')
    print(f'  Linhas de OUTRO mes ....... {len(outro_mes):>3} itens'
          f' = R$ {sum(i["valor"] for i in outro_mes):,.2f}')
    print(f'  Itens SEM VALOR ........... {len(sem_valor):>3}')
    print('=' * 68)

    graves = [l for l in todas if l['graves']]
    avisos = [l for l in todas if l['avisos'] and not l['graves']]
    print(f'  RECEITA - {len(graves)} linhas com erro grave:')
    for l in graves:
        print(f'    pasta {l["pasta"]:<6} {str(l["cliente"])[:26]:<26} {"; ".join(l["graves"])}')
    print(f'\n  RECEITA - {len(avisos)} linhas com aviso:')
    for l in avisos:
        print(f'    pasta {l["pasta"]:<6} {str(l["cliente"])[:26]:<26} {"; ".join(l["avisos"])}')
    print('=' * 68)

    # ============================================================
    # ARQUIVO DE FECHAMENTO
    # ============================================================
    out = openpyxl.Workbook()

    # ---------- RESUMO ----------
    ws = out.active
    ws.title = 'RESUMO'
    ws['A1'] = f'FECHAMENTO {args.aba_receita} - Pascoal & Dyandra Advocacia'
    ws['A1'].font = F_TIT
    ws['A2'] = f'Posicao em {hoje.strftime("%d/%m/%Y")} | origem: {args.planilha}'
    ws['A2'].font = Font(name='Calibri', italic=True, size=9)
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 52

    r = 4
    r = cabecalho(ws, r, ['RECEITA', 'VALOR', 'OBSERVACAO'])
    r = escrever(ws, r, ['Valor bruto movimentado', tot_bruto, 'inclui o que e do cliente'], moeda_cols=(2,))
    r = escrever(ws, r, ['(-) Repasses a cliente', -tot_rep, 'dinheiro de terceiro'], moeda_cols=(2,))
    r = escrever(ws, r, ['Honorario liquido lancado', tot_hon, 'soma real das linhas'], moeda_cols=(2,))
    r = escrever(ws, r, ['(-) Registro de teste', -hon_teste, 'cliente inexistente na planilha'],
                 moeda_cols=(2,), fill=P_ALERTA if hon_teste else None)
    r = escrever(ws, r, ['RECEITA REAL DO MES', tot_hon - hon_teste, ''],
                 negrito=True, moeda_cols=(2,), fill=P_OK)
    r = escrever(ws, r, ['Comissoes a pagar (I+J)', tot_com, 'a planilha nao tem esse total'], moeda_cols=(2,))

    r += 1
    r = cabecalho(ws, r, ['DESPESA', 'VALOR', 'OBSERVACAO'])
    r = escrever(ws, r, ['Pago dentro do mes', pago,
                         'so conta linha com data de pagamento no mes'], moeda_cols=(2,))
    r = escrever(ws, r, ['A pagar (ja vencido)', sum(i['valor'] for i in vencidas),
                         f'{len(vencidas)} contas vencidas'], moeda_cols=(2,),
                 fill=P_ALERTA if vencidas else None)
    r = escrever(ws, r, ['A pagar (a vencer)',
                         sum(i['valor'] for i in apagar if i not in vencidas), ''], moeda_cols=(2,))
    r = escrever(ws, r, ['Itens sem valor lancado', len(sem_valor),
                         'despesas recorrentes ainda em branco - ver aba SAIDAS'],
                 fill=P_FALTA if sem_valor else None)
    r = escrever(ws, r, ['Linhas herdadas de outro mes', len(outro_mes),
                         f'R$ {sum(i["valor"] for i in outro_mes):,.2f} - conferir se pertencem a este mes'],
                 fill=P_FALTA if outro_mes else None)

    r += 1
    r = cabecalho(ws, r, ['RESULTADO', 'VALOR', 'OBSERVACAO'])
    r = escrever(ws, r, ['Receita real', tot_hon - hon_teste, ''], moeda_cols=(2,))
    r = escrever(ws, r, ['(-) Despesa paga no mes', -pago, ''], moeda_cols=(2,))
    r = escrever(ws, r, ['(-) Despesa a pagar', -sum(i['valor'] for i in apagar), ''], moeda_cols=(2,))
    r = escrever(ws, r, ['= RESULTADO PARCIAL', tot_hon - hon_teste - pago - sum(i['valor'] for i in apagar),
                         'PARCIAL: nao inclui as despesas ainda sem valor'],
                 negrito=True, moeda_cols=(2,), fill=P_FALTA)

    # ---------- RECEITA ----------
    ws = out.create_sheet('RECEITA')
    r = cabecalho(ws, 1, ['Bloco', 'Pasta', 'Cliente', 'Empresa', 'Valor Bruto', 'Hon. Liquido',
                          'Com. I', 'Com. J', 'Repasse', 'Sobra do rateio', 'Data Repasse',
                          'PENDENCIA'],
                  [16, 9, 30, 26, 15, 15, 11, 11, 15, 15, 14, 56])
    for b in blocos:
        for l in b['linhas']:
            fill = P_ALERTA if l['graves'] else (P_FALTA if l['avisos'] else None)
            dif = l.get('diferenca')
            r = escrever(ws, r, [b['rotulo'], l['pasta'], l['cliente'], l['empresa'],
                                 num(l['bruto']), num(l['hon_liq']),
                                 num(l['com_i']), num(l['com_j']), num(l['repasse']),
                                 (dif if dif is not None and abs(dif) > TOLERANCIA else None),
                                 l['dt_repasse'], '; '.join(l['problemas'])],
                         moeda_cols=(5, 6, 7, 8, 9, 10), fill=fill)
    r = escrever(ws, r, ['TOTAL', '', '', '', tot_bruto, tot_hon,
                         sum(num(l['com_i']) or 0 for l in todas),
                         sum(num(l['com_j']) or 0 for l in todas), tot_rep, '', '', ''],
                 negrito=True, moeda_cols=(5, 6, 7, 8, 9), fill=P_OK)
    ws.freeze_panes = 'C2'
    ws.auto_filter.ref = f'A1:L{r - 1}'

    # ---------- SAIDAS ----------
    ws = out.create_sheet('SAIDAS')
    r = cabecalho(ws, 1, ['Categoria', 'Despesa', 'Valor', 'Vencimento', 'Pagamento',
                          'Situacao', 'Dias de atraso'], [26, 44, 15, 14, 14, 30, 14])
    ordem = {'A PAGAR': 0, 'SEM VALOR LANCADO': 1, 'SEM DATA': 2, 'PAGO NO MES': 3}
    for i in sorted(saidas, key=lambda x: (ordem.get(x['situacao'], 4), x['categoria'])):
        atraso = (hoje - i['venc']).days if (i['situacao'] == 'A PAGAR' and i['venc'] and i['venc'] < hoje) else None
        fill = None
        if atraso:
            fill = P_ALERTA
        elif i['situacao'] in ('SEM VALOR LANCADO', 'SEM DATA') or i['situacao'].startswith('PAGO EM'):
            fill = P_FALTA
        r = escrever(ws, r, [i['categoria'], i['desc'], i['valor'],
                             i['venc'], i['pagto'], i['situacao'], atraso],
                     moeda_cols=(3,), fill=fill)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{r - 1}'

    # ---------- PENDENCIAS ----------
    ws = out.create_sheet('PENDENCIAS')
    r = cabecalho(ws, 1, ['#', 'Gravidade', 'Onde', 'Referencia', 'O que resolver',
                          'Valor em jogo', 'Status'], [5, 12, 12, 34, 62, 16, 14])
    n = 0

    def pend(grav, onde, ref, oque, valor, fill):
        nonlocal r, n
        n += 1
        r = escrever(ws, r, [n, grav, onde, ref, oque, valor, 'A FAZER'],
                     moeda_cols=(6,), fill=fill)

    for l in sorted(todas, key=lambda x: (not x['graves'], x['pasta'])):
        ref = f'pasta {l["pasta"]} - {str(l["cliente"])[:24]}'
        if l['graves']:
            pend('GRAVE', 'RECEITA', ref, '; '.join(l['graves']),
                 num(l['hon_liq']) or num(l['bruto']), P_ALERTA)
        elif l['avisos']:
            pend('aviso', 'RECEITA', ref, '; '.join(l['avisos']),
                 num(l['hon_liq']) or num(l['bruto']), P_FALTA)

    for i in sorted(vencidas, key=lambda x: x['venc']):
        pend('GRAVE', 'SAIDAS', i['desc'][:32],
             f'vencida em {i["venc"].strftime("%d/%m/%Y")} - {(hoje - i["venc"]).days} dias de atraso',
             i['valor'], P_ALERTA)
    for i in sem_valor:
        pend('GRAVE', 'SAIDAS', i['desc'][:32],
             'despesa recorrente sem valor lancado no mes - a despesa do mes esta subavaliada',
             None, P_ALERTA)
    for i in saidas:
        if i['situacao'] == 'SEM DATA':
            pend('aviso', 'SAIDAS', i['desc'][:32], 'sem vencimento e sem pagamento',
                 i['valor'], P_FALTA)
    if outro_mes:
        pend('GRAVE', 'SAIDAS', f'{len(outro_mes)} linhas',
             'linhas com pagamento de outro mes - a aba parece copiada do mes anterior sem zerar',
             sum(i['valor'] for i in outro_mes), P_ALERTA)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:G{r - 1}'

    out.save(args.saida)
    print(f'\n  Fechamento gravado em: {args.saida}')
    print(f'  Pendencias listadas: {n}')


if __name__ == '__main__':
    main()
