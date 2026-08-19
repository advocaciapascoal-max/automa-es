"""
Provisionamento do Google Drive - Pascoal & Dyandra Advocacia.

Faz o setup inicial da integracao com o Drive da conta do escritorio:
  1. Autentica via OAuth (abre o navegador uma unica vez) e grava config/token.json
  2. Confere se a conta logada e a esperada (GOOGLE_CONTA_ESPERADA no .env)
  3. Cria a arvore de pastas do escritorio no Drive (idempotente - nao duplica)
  4. Cria a planilha de controle de contratos (.xlsx) se ainda nao existir
  5. Grava todos os IDs encontrados de volta em config/.env

Pode ser rodado quantas vezes quiser: o que ja existe e reaproveitado.

Uso:
    .venv\\Scripts\\python.exe config/setup_google_drive.py            # provisiona
    .venv\\Scripts\\python.exe config/setup_google_drive.py --dry-run  # so mostra o plano
"""
import io
import os
import sys

CONFIG_DIR = os.path.dirname(os.path.abspath(__file__))
RAIZ = os.path.join(CONFIG_DIR, '..')
sys.path.insert(0, os.path.abspath(RAIZ))

from dotenv import load_dotenv

ENV_PATH = os.path.join(CONFIG_DIR, '.env')
load_dotenv(ENV_PATH)

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaInMemoryUpload

# Mesmos escopos usados por INTEGRACOES/google_integration.py - o token gerado
# aqui e o mesmo consumido pelas automacoes depois.
SCOPES = [
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/documents',
    'https://www.googleapis.com/auth/spreadsheets'
]

MIME_PASTA = 'application/vnd.google-apps.folder'
MIME_XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

OAUTH_CREDS = os.path.join(CONFIG_DIR, 'oauth_credentials.json')
TOKEN_PATH = os.path.join(CONFIG_DIR, 'token.json')

PASTA_RAIZ = 'PASCOAL & DYANDRA - AUTOMACOES'
NOME_PLANILHA_CONTRATOS = 'Controle de Contratos.xlsx'

# Arvore de pastas: caminho relativo a raiz -> variaveis do .env que recebem o ID.
# Um mesmo ID pode alimentar mais de uma variavel (o molde tem nomes duplicados,
# ver INTAKE/google_integration.py vs INTEGRACOES/google_integration.py).
ARVORE = [
    ('CLIENTES',                        ['GOOGLE_PASTA_RECLAMANTE', 'DRIVE_PASTA_CLIENTES_ID']),
    ('MODELOS',                         []),
    ('DOCUMENTOS GERADOS',              []),
    ('DOCUMENTOS GERADOS/CONTRATOS',    ['GOOGLE_PASTA_CONTRATO_ID', 'GOOGLE_PASTA_CONTRATO']),
    ('DOCUMENTOS GERADOS/PROCURACOES',  ['GOOGLE_PASTA_PROCURACAO_ID', 'GOOGLE_PASTA_PROCURACAO']),
    ('DOCUMENTOS GERADOS/DECLARACOES',  ['GOOGLE_PASTA_DECLARACAO_ID', 'GOOGLE_PASTA_DECLARACAO']),
    ('FINANCEIRO',                      ['DRIVE_PASTA_FINANCEIRO_ID']),
    ('FINANCEIRO/FECHAMENTOS',          ['DRIVE_PASTA_FECHAMENTO_ID']),
    ('CONTROLE',                        []),
]


# ---------------------------------------------------------------- autenticacao

def autenticar():
    """OAuth com a conta do escritorio. Reaproveita config/token.json se existir."""
    if not os.path.exists(OAUTH_CREDS):
        print('ERRO: config/oauth_credentials.json nao encontrado.')
        print()
        print('Crie o OAuth Client ID (tipo "Desktop app") no Google Cloud Console,')
        print('baixe o JSON e salve exatamente como:')
        print(f'   {OAUTH_CREDS}')
        print()
        print('Passo a passo em docs/SETUP_GOOGLE_DRIVE.md')
        sys.exit(1)

    creds = None
    if os.path.exists(TOKEN_PATH):
        try:
            creds = Credentials.from_authorized_user_file(TOKEN_PATH, SCOPES)
        except Exception:
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print('Renovando token de acesso...')
            creds.refresh(Request())
        else:
            print('Abrindo o navegador para login na conta do escritorio...')
            print('(so precisa ser feito uma vez)')
            flow = InstalledAppFlow.from_client_secrets_file(OAUTH_CREDS, SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_PATH, 'w') as fh:
            fh.write(creds.to_json())
        print(f'Token salvo em {TOKEN_PATH}')

    drive = build('drive', 'v3', credentials=creds)
    return drive


def conferir_conta(drive):
    """Mostra a conta logada e alerta se nao for a esperada."""
    sobre = drive.about().get(fields='user,storageQuota').execute()
    email = sobre['user'].get('emailAddress', '?')
    esperada = os.getenv('GOOGLE_CONTA_ESPERADA', '').strip()

    quota = sobre.get('storageQuota', {})
    limite = int(quota.get('limit', 0) or 0)
    usado = int(quota.get('usage', 0) or 0)

    print(f'Conta logada: {email}')
    if limite:
        print(f'Armazenamento: {usado / 2**30:.2f} GB de {limite / 2**30:.0f} GB')

    if esperada and email.lower() != esperada.lower():
        print()
        print(f'ATENCAO: a conta esperada e {esperada}, mas quem logou foi {email}.')
        print('Se estiver errado, apague config/token.json e rode de novo.')
        resp = input('Continuar mesmo assim? [s/N] ').strip().lower()
        if resp != 's':
            sys.exit(1)
    return email


# ------------------------------------------------------------------- utilidades

def _escapar(nome):
    """Escapa aspas simples para a query do Drive."""
    return nome.replace('\\', '\\\\').replace("'", "\\'")


def achar_pasta(drive, nome, pai_id=None):
    """Procura uma pasta pelo nome (dentro do pai, se informado). Retorna o ID ou None."""
    query = (
        f"name = '{_escapar(nome)}' and mimeType = '{MIME_PASTA}' and trashed = false"
    )
    query += f" and '{pai_id}' in parents" if pai_id else " and 'root' in parents"

    res = drive.files().list(
        q=query, fields='files(id, name)', pageSize=10,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    arquivos = res.get('files', [])
    return arquivos[0]['id'] if arquivos else None


def garantir_pasta(drive, nome, pai_id=None, dry_run=False):
    """Retorna o ID da pasta, criando-a se ainda nao existir. Idempotente."""
    existente = achar_pasta(drive, nome, pai_id)
    if existente:
        print(f'   [ja existe] {nome}  ({existente})')
        return existente

    if dry_run:
        print(f'   [criaria  ] {nome}')
        return None

    corpo = {'name': nome, 'mimeType': MIME_PASTA}
    if pai_id:
        corpo['parents'] = [pai_id]

    pasta = drive.files().create(
        body=corpo, fields='id', supportsAllDrives=True
    ).execute()
    print(f'   [criada   ] {nome}  ({pasta["id"]})')
    return pasta['id']


def achar_arquivo(drive, nome, pai_id):
    query = (
        f"name = '{_escapar(nome)}' and trashed = false and '{pai_id}' in parents"
    )
    res = drive.files().list(
        q=query, fields='files(id, name)', pageSize=10,
        supportsAllDrives=True, includeItemsFromAllDrives=True,
    ).execute()
    arquivos = res.get('files', [])
    return arquivos[0]['id'] if arquivos else None


# ------------------------------------------------ planilha de controle de contratos

def montar_planilha_contratos(ano):
    """
    Gera o .xlsx de controle de contratos no layout que
    INTEGRACOES/google_integration.py:buscar_proximo_contrato() espera:
      - aba 'Contratos {ano}'
      - cabecalho na linha 4, dados a partir da linha 5
      - col D = Nome | F = Data cadastro | H = CT. | I = Parte contraria | J = Area
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Contratos {ano}'

    ws['B2'] = f'CONTROLE DE CONTRATOS - {ano}'
    ws['B2'].font = Font(bold=True, size=14)

    cabecalhos = {
        'D': 'NOME DO CLIENTE',
        'F': 'DATA DE CADASTRO',
        'H': 'CT.',
        'I': 'PARTE CONTRARIA',
        'J': 'AREA',
    }
    for col, titulo in cabecalhos.items():
        celula = ws[f'{col}4']
        celula.value = titulo
        celula.font = Font(bold=True)
        celula.alignment = Alignment(horizontal='center')

    larguras = {'B': 4, 'D': 38, 'F': 18, 'H': 12, 'I': 34, 'J': 18}
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura

    # Linha-semente: buscar_proximo_contrato() varre a coluna H procurando
    # "N/{ano}" para descobrir o ultimo numero. Sem nenhuma linha, ele comeca
    # do 1 normalmente - a semente abaixo so documenta o formato esperado.
    ws['H5'] = f'0/{ano}'
    ws['D5'] = '(linha-semente - nao apagar o formato da coluna H)'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def garantir_planilha_contratos(drive, pasta_controle_id, ano, dry_run=False):
    if not pasta_controle_id:
        return None

    existente = achar_arquivo(drive, NOME_PLANILHA_CONTRATOS, pasta_controle_id)
    if existente:
        print(f'   [ja existe] {NOME_PLANILHA_CONTRATOS}  ({existente})')
        return existente

    if dry_run:
        print(f'   [criaria  ] {NOME_PLANILHA_CONTRATOS}')
        return None

    conteudo = montar_planilha_contratos(ano)
    # Sobe como .xlsx binario de proposito: buscar_proximo_contrato() usa
    # files().get_media(), que NAO funciona em planilha nativa do Google.
    media = MediaInMemoryUpload(conteudo, mimetype=MIME_XLSX, resumable=False)
    arquivo = drive.files().create(
        body={'name': NOME_PLANILHA_CONTRATOS, 'parents': [pasta_controle_id]},
        media_body=media, fields='id', supportsAllDrives=True,
    ).execute()
    print(f'   [criada   ] {NOME_PLANILHA_CONTRATOS}  ({arquivo["id"]})')
    return arquivo['id']


# ------------------------------------------------------------------ escrita .env

def gravar_env(valores):
    """
    Atualiza as chaves indicadas em config/.env preservando comentarios e ordem.
    Chaves que nao existirem no arquivo sao acrescentadas ao final.
    """
    if not os.path.exists(ENV_PATH):
        print(f'ERRO: {ENV_PATH} nao existe. Copie de config/.env.example primeiro.')
        sys.exit(1)

    with open(ENV_PATH, 'r', encoding='utf-8') as fh:
        linhas = fh.readlines()

    pendentes = dict(valores)
    saida = []
    for linha in linhas:
        despido = linha.strip()
        if despido and not despido.startswith('#') and '=' in despido:
            chave = despido.split('=', 1)[0].strip()
            if chave in pendentes:
                saida.append(f'{chave}={pendentes.pop(chave)}\n')
                continue
        saida.append(linha)

    if pendentes:
        saida.append('\n# Adicionado por config/setup_google_drive.py\n')
        for chave, valor in pendentes.items():
            saida.append(f'{chave}={valor}\n')

    # Backup antes de sobrescrever (config/*.bak ja esta no .gitignore).
    backup = ENV_PATH + '.bak'
    with open(backup, 'w', encoding='utf-8') as fh:
        fh.writelines(linhas)

    with open(ENV_PATH, 'w', encoding='utf-8') as fh:
        fh.writelines(saida)

    print(f'\nconfig/.env atualizado ({len(valores)} variaveis). Backup: config/.env.bak')


# ------------------------------------------------------------------------ main

def main():
    dry_run = '--dry-run' in sys.argv
    from datetime import datetime
    ano = datetime.now().year

    print('=' * 68)
    print('  PROVISIONAMENTO DO GOOGLE DRIVE - Pascoal & Dyandra Advocacia')
    if dry_run:
        print('  MODO DRY-RUN - nada sera criado')
    print('=' * 68)
    print()

    drive = autenticar()
    conferir_conta(drive)
    print()

    print(f'Arvore de pastas em "{PASTA_RAIZ}":')
    raiz_id = garantir_pasta(drive, PASTA_RAIZ, None, dry_run)

    ids_por_caminho = {'': raiz_id}
    valores_env = {}

    for caminho, chaves_env in ARVORE:
        partes = caminho.split('/')
        nome = partes[-1]
        caminho_pai = '/'.join(partes[:-1])
        pai_id = ids_por_caminho.get(caminho_pai)

        if pai_id is None and not dry_run:
            print(f'   [pulado   ] {caminho} (pasta pai nao resolvida)')
            continue

        pasta_id = garantir_pasta(drive, nome, pai_id, dry_run)
        ids_por_caminho[caminho] = pasta_id

        if pasta_id:
            for chave in chaves_env:
                valores_env[chave] = pasta_id

    print()
    print('Planilha de controle:')
    planilha_id = garantir_planilha_contratos(
        drive, ids_por_caminho.get('CONTROLE'), ano, dry_run)
    if planilha_id:
        valores_env['GOOGLE_SHEETS_CONTRATOS_ID'] = planilha_id
        valores_env['GOOGLE_PLANILHA_CONTRATOS'] = planilha_id
        valores_env['GOOGLE_ABA_CONTRATOS'] = f'Contratos {ano}'

    if dry_run:
        print('\nDRY-RUN: nada foi criado e o .env nao foi tocado.')
        return

    if valores_env:
        gravar_env(valores_env)

    if raiz_id:
        print(f'\nPasta raiz no Drive:')
        print(f'   https://drive.google.com/drive/folders/{raiz_id}')

    print()
    print('-' * 68)
    print('PENDENTE (nao da pra automatizar - depende de material do escritorio):')
    print('  * GOOGLE_TEMPLATE_ID ............ Google Doc da Ficha do Cliente')
    print('  * GOOGLE_TEMPLATE_CONTRATO_ID ... Google Doc do Contrato de Honorarios')
    print('  * GOOGLE_TEMPLATE_PROCURACAO_ID . Google Doc da Procuracao')
    print('  * GOOGLE_TEMPLATE_DECLARACAO_ID . Google Doc da Declaracao de Hipossuficiencia')
    print()
    print('  Suba os 4 modelos na pasta MODELOS do Drive, com os placeholders')
    print('  {{Nome do cliente}}, {{CPF}}, etc. (ver mapa_chaves em')
    print('  INTEGRACOES/google_integration.py) e rode:')
    print('     .venv\\Scripts\\python.exe config/vincular_modelos.py')
    print('-' * 68)


if __name__ == '__main__':
    try:
        main()
    except HttpError as erro:
        print(f'\nERRO da API do Google: {erro}')
        sys.exit(1)
    except KeyboardInterrupt:
        print('\nCancelado.')
        sys.exit(1)
